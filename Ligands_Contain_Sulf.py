import os
import gzip
import re
from tqdm import tqdm
from openpyxl import Workbook
from concurrent.futures import ThreadPoolExecutor, as_completed

EXCLUDE_LIST = []
MODRES_LIGANDS = []

def contains_sulfur(formula):
    elements = formula.split()
    sulfur_pattern = r'(^S$|S\d+)'
    for element in elements:
        if re.search(sulfur_pattern, element):
            return True
    return False

def should_exclude(ligand_name):
    return any(exclude in ligand_name.upper() for exclude in EXCLUDE_LIST)

def process_single_pdb_file(pdb_file, ligands_count_dict):
    global MODRES_LIGANDS
    pdb_id = pdb_file.split('.')[0]
    ligands = []
    MODRES_LIGANDS = []

    try:
        with gzip.open(pdb_file, 'rt') as f:
            for line in f:
                if line.startswith('MODRES'):
                    parts = line.split()
                    if len(parts) >= 3:
                        modres_ligand_name = parts[2]
                        MODRES_LIGANDS.append(modres_ligand_name)
                if line.startswith('FORMUL'):
                    parts = line.split()
                    if len(parts) >= 4:
                        ligand_name = parts[2]
                        formula_match = re.search(r'\((.*?)\)', line)
                        if formula_match:
                            formula = formula_match.group(1)
                        else:
                            formula = " ".join(parts[3:])
                        if contains_sulfur(formula):
                            if not should_exclude(ligand_name) and ligand_name not in MODRES_LIGANDS:
                                ligands.append(f"{ligand_name}\t({formula})\t")
                                if ligand_name in ligands_count_dict:
                                    ligands_count_dict[ligand_name] += 1
                                else:
                                    ligands_count_dict[ligand_name] = 1
    except (IOError, EOFError) as e:
        print(f"Error opening or reading compressed file {pdb_file}: {e}, skipping this file.")

    return pdb_id, ligands

def process_pdb_files(output_dir):
    pdb_files = [f for f in os.listdir('.') if f.endswith('.pdb.gz')]
    ligands_count_dict = {}

    output_file = os.path.join(output_dir, "sulfur_containing_ligands.txt")
    if not os.path.exists(output_dir):
        try:
            os.makedirs(output_dir)
        except OSError as e:
            print(f"Error creating directory {output_dir}: {e}, the program may not continue normally, please check.")
            return

    try:
        with open(output_file, 'w') as out_f, ThreadPoolExecutor(max_workers=3) as executor:
            future_to_file = {executor.submit(process_single_pdb_file, pdb_file, ligands_count_dict): pdb_file for pdb_file in pdb_files}

            for future in tqdm(as_completed(future_to_file), total=len(future_to_file), desc="Processing PDB files"):
                pdb_file = future_to_file[future]
                try:
                    pdb_id, ligands = future.result()
                    if ligands:
                        out_f.write(f"{pdb_id}\t" + "\t".join(ligands) + "\n")
                except Exception as e:
                    print(f"Error processing file {pdb_file}: {e}")

            write_to_excel(ligands_count_dict, output_dir)
    except IOError as e:
        print(f"Error writing to text file {output_file}: {e}, statistical results may not be saved correctly, please check.")

def write_to_excel(ligands_count_dict, output_dir):
    excel_file_path = os.path.join(output_dir, "ligands_count.xlsx")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ligands Count"

        ws["A1"] = "Ligand Name"
        ws["B1"] = "Count"

        sorted_ligands = sorted(ligands_count_dict.items(), key=lambda x: x[1], reverse=True)

        for row, (ligand_name, count) in enumerate(sorted_ligands, start=2):
            ws.cell(row=row, column=1, value=ligand_name)
            ws.cell(row=row, column=2, value=count)

        wb.save(excel_file_path)
    except IOError as e:
        print(f"Error writing to Excel file {excel_file_path}: {e}, please check permissions or if the file is in use.")

if __name__ == "__main__":
    output_dir = "/data0/caiwenhao/PDB"
    process_pdb_files(output_dir)
    print(f"Processing complete. Results have been saved to sulfur_containing_ligands.txt and ligands_count.xlsx in the {output_dir} directory.")